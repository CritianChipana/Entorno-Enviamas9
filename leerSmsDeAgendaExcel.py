# -*- coding: utf-8 -*-
"""
Created on Wed Apr  6 10:06:17 2022

@author: Cristian

"""
from decouple import config
# import traceback
import pymysql
from datetime import datetime
import openpyxl
import requests
import json
from math import ceil

class Model:

    new_url = None

    def __init__(self):
        self.connection = pymysql.connect(
            host=config('MYSQL_HOST'),
            user=config('MYSQL_USER'),
            password=config('MYSQL_PASSWORD'),
            db=config('MYSQL_DB')
        )

        self.cursor = self.connection.cursor()
        print('coneccion estableciada correctamente')

    def select_all_campaign_agenda_excel_model(self):
        # pendiente 2 y agendado 4 el status de cada campaña
        sql = 'SELECT * FROM campaigns where (campaign_type_id = 1 || campaign_type_id = 2 || campaign_type_id = 3 || campaign_type_id = 4) and service_id = 1 and (status = 2 || status = 4) order by scheduled asc '
        # sql = 'SELECT * FROM campaigns where id=9'
        try:
            self.cursor.execute(sql)
            campaigns = self.cursor.fetchone()
            return campaigns
        except Exception as e:
            print(e)
            raise

    def select_all_contact_by_agenda_model(self, id):
        sql = 'SELECT * FROM contacts WHERE agenda_id = {} and state = true '.format(id)
        try:
            self.cursor.execute(sql)
            campaigns = self.cursor.fetchall()
            return campaigns
        except Exception as e:
            print(e)
            raise

    def select_sms_campaign_by_id(self, id):
        sql = 'SELECT * FROM sms_campaigns WHERE campaign_id = {} '.format(id)
        try:
            self.cursor.execute(sql)
            campaigns = self.cursor.fetchone()
            return campaigns
        except Exception as e:
            print(e)
            raise

    def crear_sms(self, payload):
        print('rrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrr')
        sql = "INSERT INTO sms (credit, is_push, content, phone, status, commit, response, message_id ,payload ,user_id, campaign_id, channel_id,created_at,updated_at, send_at) VALUES ({},{},'{}','{}','{}','{}','{}',{},'{}',{},{},{},'{}','{}', '{}')".format(
            payload['credit'], payload['is_push'], payload['content'], payload['phone'], payload['status'], payload['commit'], payload['response'], payload['message_id'],payload['payload'],payload['user_id'], payload['campaign_id'], payload['channel_id'],payload['created_at'], payload['updated_at'], payload['send_at'])
        sql2 = "SELECT LAST_INSERT_ID()"
        try:
            self.cursor.execute(sql)
            self.connection.commit()
            self.cursor.execute(sql2)
            sms_id = self.cursor.fetchone()
            print('se creo sms')
            return sms_id[0]
        except Exception as e:
            print(e)
            return False

    def crear_sms_with_message_id(self, payload):
        print('rrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrr')
        sql = "INSERT INTO sms (credit, is_push, content, phone, status, commit, response, message_id ,payload ,user_id, campaign_id, channel_id,created_at,updated_at, send_at, message_id_external) VALUES ({},{},'{}','{}','{}','{}','{}',{},'{}',{},{},{},'{}','{}', '{}', {})".format(
            payload['credit'], payload['is_push'], payload['content'], payload['phone'], payload['status'], payload['commit'], payload['response'], payload['message_id'],payload['payload'],payload['user_id'], payload['campaign_id'], payload['channel_id'],payload['created_at'], payload['updated_at'], payload['send_at'], payload['message_id'])
        sql2 = "SELECT LAST_INSERT_ID()"
        try:
            self.cursor.execute(sql)
            self.connection.commit()
            self.cursor.execute(sql2)
            sms_id = self.cursor.fetchone()
            print('se creo sms')
            return sms_id[0]
        except Exception as e:
            print(e)
            return False

    def select_excel_sms_by_id (self,id):
        sql = 'SELECT * FROM excel_sms_campaigns WHERE campaign_id = {} '.format(id)
        try:
            self.cursor.execute(sql)
            excel_sms_campaigns = self.cursor.fetchone()
            return excel_sms_campaigns
        except Exception as e:
            print(e)
            raise

    def select_json_sms_by_id (self,id):
        sql = 'SELECT * FROM json_sms_campaigns WHERE campaign_id = {} '.format(id)
        try:
            self.cursor.execute(sql)
            excel_sms_campaigns = self.cursor.fetchone()
            return excel_sms_campaigns
        except Exception as e:
            print(e)
            raise

    def change_state_campaign(self, campaign_id, status):
        sql = "UPDATE campaigns SET status = {} where id = {}".format(
            status, campaign_id)
        try:
            self.cursor.execute(sql)
            self.connection.commit()
            print('se modificao el status de la campaña ' + "{}".format(status))
            return True
        except Exception as e:
            print(e)
            return False

    def select_user(self, user_id):
        sql = "SELECT * FROM users where id ={} ".format(user_id)
        try:
            self.cursor.execute(sql)
            user = self.cursor.fetchone()
            return user
        except Exception as e:
            print(e)
            return False

    def create_url(self, data):
        print('crear url')
        sql = "INSERT INTO `urls`(`name`, `short_url`, `long_url`, `user_id`, `group_url_id`, `url_id`, `status`, `state`, `created_at`, `updated_at`) VALUES ('{}','{}','{}',{},{},'{}',{},{},'{}','{}')".format(data['name'],data['short_url'],data['long_url'],data['user_id'],data['group_url_id'],data['url_id'],data['status'],data['state'],data['created_at'],data['updated_at'],)
        sql2 = "SELECT LAST_INSERT_ID()"

        try:
            self.cursor.execute(sql)
            self.connection.commit()
            self.cursor.execute(sql2)
            new_url_id = self.cursor.fetchone()
            return new_url_id[0]
        except Exception as e:
            print(e)
            return False

    def select_group_url(self, campaign_id):
        sql = "SELECT * FROM group_urls where campaign_id = {} ".format(campaign_id)
        try:
            self.cursor.execute(sql)
            group_url = self.cursor.fetchone()
            return group_url
        except Exception as e:
            print(e)
            return False

    def select_channel_by_id(self, channel_id):
        sql = "SELECT * FROM channels where id = {} ".format(channel_id)
        
        print('holis')
        print(channel_id)

        try:
            self.cursor.execute(sql)
            channel = self.cursor.fetchone()
            return channel
        except Exception as e:
            print(e)
            return False

    def select_user_channels(self, user_id):
        sql = "select c.*  from channel_user cu , users u , channels c where cu.user_id = u.id and c.id = cu.channel_id and u.id = {} ".format(user_id)

        try:
            self.cursor.execute(sql)
            channels = self.cursor.fetchall()
            return channels
        except Exception as e:
            print(e)
            return False

    def send_sms_to_provider_by_id(self, provider_id):
        sql = "SELECT * FROM providers where id = {} ".format(provider_id)
        try:
            self.cursor.execute(sql)
            provider = self.cursor.fetchone()
            return provider
        except Exception as e:
            print(e)
            return False

    def create_sms_url(self, data):
        print('crear url')
        sql = "INSERT INTO `sms_url`(`sms_id`, `url_id`, `created_at`, `updated_at`) VALUES ({},{},'{}','{}')".format(data['sms_id'],data['url_id'],data['created_at'],data['updated_at'],)

        try:
            self.cursor.execute(sql)
            self.connection.commit()
        
            return True
        except Exception as e:
            print(e)
            return False

    def close(self):
        self.connection.close()

class Controller(object):

    def __init__(selft):
        selft.model = Model()
        selft.view = View()

    def send_sms_by_agenda(self, campaign, contacts_by_agenda, sms_campaign):

        user_id = campaign[5]

        user = self.model.select_user(user_id)

        user_chanel_id = user[7]
        user_email = user[2]
        
        f.write('\n' + 'El usuario usa el channel: ' + str(user_chanel_id))
        f.write('\n' + 'El email del usuario es: ' + str(user_email))
        f.write('\n' + '*********** Inicio de envio de sms ******************')

        key_contacts = [
            {
                "key": "[NOMBRE1]",
                "value": 2,
            },
            {
                "key": "[NOMBRE2]",
                "value": 3,
            },
            {
                "key": "[APELLIDO1]",
                "value": 5,
            },
            {
                "key": "[APELLIDO2]",
                "value": 6,
            },
            {
                "key": "[EMAIL]",
                "value": 4,
            },
            {
                "key": "[VAR1]",
                "value": 7,
            },
            {
                "key": "[VAR2]",
                "value": 8,
            },
            {
                "key": "[VAR3]",
                "value": 9,
            },
            {
                "key": "[VAR4]",
                "value": 10,
            },
        ]
        
        contador_de_sms_enviados = 0

        for contact in contacts_by_agenda:
            contador_de_sms_enviados = contador_de_sms_enviados + 1
            f.write('\n' + 'Sms '+ str(contador_de_sms_enviados) + ' de '+ str(len(contacts_by_agenda)))

            message = sms_campaign[2]

            for clave in key_contacts:
                if(clave['key'] in message and contact[clave['value']] != None) :
                    message = message.replace(clave['key'], contact[clave['value']])  
                else:
                    message=message.replace(clave['key'], "")

            # usar url individual
            message = self.has_individual_url(sms_campaign,campaign,message)

            if( message == None ):
                exit()
            

            message = self.standardize_message(message)
            number = contact[1]
            credit = self.calculate_credits(message)

            try:
                phone_status = self.validate_phone(int(number))
            except ValueError:
                phone_status = False
                pass    


            # seleccionar el proveedor
            if(phone_status):
                response = self.send_sms_to_provider(user_chanel_id, message, number, campaign, sms_campaign)
                response = list(response)
                if(response[2] != 0):
                    response.append('DELIVERED')
                else:
                    response.append('REJECTED')
            else: 
                response = ('a','b',0,'REJECTED')

            # mandar la informacion para crear el sms
            result = self.send_sms(credit,sms_campaign,message,contact[1] , response, campaign,user, None)
            print('///////////////////////////////////')


        print(contact[0])

    def send_sms_individuales(self, campaign, contacts, sms_campaign):

        user_id = campaign[5]

        user = self.model.select_user(user_id)

        user_chanel_id = user[7]
        user_email = user[2]
        
        f.write('\n' + 'El usuario usa el channel: ' + str(user_chanel_id))
        f.write('\n' + 'El email del usuario es: ' + str(user_email))
        f.write('\n' + '*********** Inicio de envio de sms ******************')

      
        
        contador_de_sms_enviados = 0

        for contact in contacts:
            contador_de_sms_enviados = contador_de_sms_enviados + 1
            f.write('\n' + 'Sms '+ str(contador_de_sms_enviados) + ' de '+ str(len(contacts)))

            message = sms_campaign[2]

            # usar url individual
            message = self.has_individual_url(sms_campaign,campaign,message)

            if( message == None ):
                exit()
            

            message = self.standardize_message(message)
            number = contact
            credit = self.calculate_credits(message)

            try:
                phone_status = self.validate_phone(int(number))
            except ValueError:
                phone_status = False
                pass    


            # seleccionar el proveedor
            if(phone_status):
                response = self.send_sms_to_provider(user_chanel_id, message, number, campaign, sms_campaign)
                response = list(response)
                if(response[2] != 0):
                    response.append('DELIVERED')
                else:
                    response.append('REJECTED')
            else: 
                response = ('a','b',0,'REJECTED')

            # mandar la informacion para crear el sms
            result = self.send_sms(credit,sms_campaign,message,contact , response, campaign,user, None)
            print('///////////////////////////////////')

    def read_excel(self, path, sms_campaign, campaign):

        user = self.model.select_user(campaign[5])

        f.write('\n' + 'El usuario usa el channel: ' + str(user[7]))
        user_chanel_id = user[7]

        f.write('\n' + 'El email del usuario es: ' + str(user[2]))
        f.write('\n' + '*********** Inicio de envio de sms ******************')
        # leer el archivo opcion 1
        key_contacts = [
            {
                "key": "[VAR1]",
                "value": 2,
            },
            {
                "key": "[VAR2]",
                "value": 3,
            },
            {
                "key": "[VAR3]",
                "value": 4,
            },
            {
                "key": "[VAR4]",
                "value": 5,
            },
            {
                "key": "[VAR5]",
                "value": 6,
            },
            {
                "key": "[VAR6]",
                "value": 7,
            },
            {
                "key": "[VAR7]",
                "value": 8,
            },
            {
                "key": "[VAR8]",
                "value": 9,
            }
        ]

        print('excel ***********')
        # print(path)
        book = openpyxl.load_workbook(path, data_only=True)
        sheet_name = book[config('NAME_OF_ACTIVE_SHEET')]
        row_count = sheet_name.max_row
        print("###########")

        for i in range(1,row_count):
            message = sms_campaign[2]

            for clave in key_contacts:
                if sheet_name.cell(row = i+1, column = 1).value != None:
                    phone = (sheet_name.cell(row = i+1, column = 1).value)
                    print(phone)

                if(clave['key'] in message and sheet_name.cell(row = i+1, column = clave['value']).value != None ):
                    message = message.replace(clave['key'], str(sheet_name.cell(row = i+1, column=clave['value']).value))
                else:
                    message=message.replace(clave['key'], "")

            if sheet_name.cell(row=i+1, column=1).value !=None:
                
                message = self.standardize_message(message)
                #Crear url individual
                message=  self.has_individual_url(sms_campaign, campaign, message)

                phone_status = self.validate_phone(phone)

                credit = self.calculate_credits(message)

                # seleccionar el proveedor
                if(phone_status):
                    response = self.send_sms_to_provider(user_chanel_id, message, phone, campaign, sms_campaign)
                    response = list(response)
                    if(response[2] != 0):
                        response.append('DELIVERED')
                    else:
                        response.append('REJECTED')
                else: 
                    response = ('no se envio mensaje por el status del numero','b',0,'REJECTED')

                #Mandar datos para crear sms
                result = self.send_sms(credit,sms_campaign,message,phone ,response, campaign,user, None)

                print("cccccccccccccccccccccccccccccccccccccc")
    
    def read_json(self, path, sms_campaign, campaign):
        
        try:
            
            user = self.model.select_user(campaign[5])
            self.new_url = None

            f.write('\n' + 'El usuario usa el channel: ' + str(user[7]))
            user_chanel_id = user[7]
            f.write('\n' + 'El email del usuario es: ' + str(user[2]))
            f.write('\n' + '*********** Inicio de envio de sms ******************')
            with open(path, 'r') as fe:
                data = json.load(fe)
            print(7)
            print(data[0]['phone'])

            for i in range(0,len(data)):
                if "phone" in data[i]:
                    phone = data[i]['phone']
                else: 
                    phone = None

                if "text" in data[i]:
                    message = data[i]['text']
                else: 
                    message = ""

                if "message_id" in data[i]:
                    message_id = data[i]['message_id']
                else: 
                    message_id = None

                message = self.standardize_message(message)

                phone_status = self.validate_phone(phone)

                credit = self.calculate_credits(message)

                # seleccionar el proveedor
                if(phone_status):
                    response = self.send_sms_to_provider(user_chanel_id, message, phone, campaign, sms_campaign)
                    response = list(response)
                    if(response[2] != 0):
                        response.append('DELIVERED')
                    else:
                        response.append('REJECTED')
                else: 
                    response = ('no se envio mensaje por el status del numero','b',0,'REJECTED')

                #Mandar datos para crear sms
                result = self.send_sms(credit,sms_campaign,message,phone ,response, campaign,user, message_id)

                print("cccccccccccccccccccccccccccccccccccccc")  
        except KeyError:
            f.write('No se pudo enviar sms', data)
            print('Nooooooooooooooooooooooooooooo')

    def send_sms(self,credit, sms_campaign, message, phone, response ,campaign,user, message_id):
        print('se envio sms')

        channelsUser = self.model.select_user_channels(campaign[5])

        # preguntar si la campaña fue lanzado ccomo bidireccional
        channel_selected = 0
        if not sms_campaign[1]:
            channel_selected = channelsUser[0][0]
            print("♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠")
        else:
            channel_selected = channelsUser[1][0]
            print("☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻")

        print("???????????????????????????????????????????????????????????????????????")
        print(channel_selected)
        print("???????????????????????????????????????????????????????????????????????")

        if  response[3] == "REJECTED":
            credit = 0

        payload = {
                    'credit': credit,
                    "is_push": sms_campaign[3],
                    "content": message,
                    "phone": phone,
                    "status": response[3],
                    "commit": "",
                    "response": response[1],
                    "message_id": response[2],
                    "payload": str(response[0]),
                    "user_id": campaign[5],
                    "campaign_id": sms_campaign[7],
                    "channel_id": channel_selected,
                    "created_at": datetime.now(),
                    "updated_at":datetime.now(),
                    "send_at":datetime.now()
                }
        if message_id != None: 
            payload['message_id'] = message_id

        f.write('\n' + 'hora de envio de sms:' + str(datetime.now()))
        f.write('\n' + 'datos del sms:')
        f.write(f'\n {repr(payload)}')
        f.write(f'\n')
        
        
        if message_id != None:
            sms_id = self.model.crear_sms_with_message_id(payload)
        else:
            sms_id = self.model.crear_sms(payload)

        print('7777777777777777')
        print(sms_id)
        if(self.new_url):
            #CREAR TABA PARA RELACIONAR SMS CON URL
            payload_url = {
                    'sms_id': sms_id,
                    'url_id': self.new_url,
                    'created_at':datetime.now(),
                    'updated_at':datetime.now(),
                }
            self.new_url = self.model.create_sms_url(payload_url)


            print(self.new_url)
        print('7777777777777777')
        return sms_id

    def create_cut_url(self, long_url):
        payload = json.dumps({
        "url_register": long_url,
        "type": 2
        })
        
        headers = {
            'Authorization': 'Basic YXBwQGVudmlhbWFzLnBlOkRldmVsb3BtZW50JCQyMDIy=',
            'Content-Type': 'application/json'
        }

        response = requests.post(config('ENDPOINT_CUT_PE'), headers=headers, data=payload)
        dataJson = response.json()

        print((  dataJson))
        print(dataJson['data']['shortUrl'])
        return dataJson['data']

    def calculate_credits(self, message):
        print('calcular mensage')
        n=0
        m=0
        for f in message:
            if  message[n]== '|' or message[n]== '{' or message[n]== '}' or message[n]== '[' or message[n]== ']':
                m = m + 1
            n = n + 1
            m = m + 1

        if m <= 160:
            per_message = 160
        else:
            per_message = 153
        credit = int(ceil(m / float(per_message)))
        return credit

    def standardize_message(self, message):
        print('mensaje estandarizado')
        message=message.replace("'", "")
        message=message.replace("ñ", "n")
        message=message.replace("Ñ", "N")
        message=message.replace("á", "a")
        message=message.replace("Á", "A")
        message=message.replace("é", "e")
        message=message.replace("É", "E")
        message=message.replace("í", "i")
        message=message.replace("Í", "I")
        message=message.replace("ó", "o")
        message=message.replace("Ó", "O")
        message=message.replace("ú", "u")
        message=message.replace("Ú", "U")
        return message

    def validate_phone(self, phone):
        try:
          
            print('validacion de celular')
            phone = int(phone)

            if( len(str(phone)) == 9  and type(phone) == int):
                print('Numero valido')
                return True
            else:
                return False
        except Exception as e: 
                return False
               

    def has_individual_url(self, sms_campaign, campaign, auxiliar):
        
        if (sms_campaign[4] == None and sms_campaign[5]!= None and ('[CUSTOM_URL]' in auxiliar)):
            print('tiene link corto')
            f.write('\n' + 'Tiene link corto personalizado')

            try:
                group_url = self.model.select_group_url(campaign[0])
                url = self.create_cut_url(sms_campaign[5])
                auxiliar = auxiliar.replace( '[CUSTOM_URL]', url['shortUrl'] )

                payload_url = {
                    'name':'URL individual',
                    'short_url':url['shortUrl'],
                    'long_url':sms_campaign[5],
                    "user_id": campaign[5],
                    'group_url_id':group_url[0],
                    'url_id':url['url_id'],
                    'status':False,
                    'state':True,
                    'created_at':datetime.now(),
                    'updated_at':datetime.now(),
                }
                self.new_url = self.model.create_url(payload_url)

                return auxiliar

            except Exception as e: 
                f.write('\n' + 'ERROR:!!!!!' + 'No se pudo crear url personalizado, verifique credenciales o el path de cut.pe')
                f.write('\n' + str(e))
                
                print("tttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttt")
                print("Ocurrio un Problema a la hora de cortar url, No se pudo crear url personalizado")
                print(str(e))
                print("tttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttt")
                # exit()
                # sys.exit()
                # quit()
                return None

        else: 
            self.new_url = None
            return auxiliar

    def send_sms_to_provider(self, channel_id, message, number, campaign, sms_campaign):

        print('channel_id')
        f.write('\n' + 'Enviando por el canal con id: ' + str(channel_id))
        user_id = campaign[5]
        # listar los channesl del usuario
        channelsUser = self.model.select_user_channels(user_id)

        print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        print(channelsUser)
        print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        print(channelsUser[0][0])
        print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")

        # preguntar si la campaña fue lanzado ccomo bidireccional
        channel_selected = 0
        if not sms_campaign[1]:
            channel_selected = channelsUser[0][0]
            print("♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠♠")
        else:
            channel_selected = channelsUser[1][0]
            print("☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻☻")

        print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        print(channel_selected)
        print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")

        channel = self.model.select_channel_by_id(channel_selected)

        provider_id = channel[11]
        api_key = channel[5]
        dial = channel[6]
        authorization = channel[7]
        mask = channel[14]

        f.write('\n' + 'Enviando por el provider con id: ' + str(provider_id))

        if( provider_id == 1 ):
# "Es bonito" if es_bonito else "No es bonito"
            try:
                payload = json.dumps({
                    "apiKey": api_key,
                    "country": "PE",
                    "dial": dial,
                    "message": message,
                    "msisdns": [
                        '51' + str(number)
                    ],
                    "tag": "ENVIAMAS2_" + str(campaign[0]),
                    "mask": mask,
                    "dlr": True,
                    "msgClass": 0 if sms_campaign[3] == 1 else 1,
                    "optionals": "{registeredDelivery:11}"
                })

                headers = {
                    'Authorization': str(authorization),
                    'Content-Type': 'application/json'
                }
                response = requests.request("POST", config('ENDPOINT_PROVEEDOR'), headers = headers, data=payload)
                dataJson = response.json()
                print(1111111111111111111111111)
                print(dataJson)
                print(1111111111111111111111111)
                mailingId = dataJson['mailingId']
                print("Respuesta de proveedor:")
                print(response.text)
                data_text = response.text

                return json.dumps(payload), data_text, mailingId

            except Exception as e: 
                f.write('\n' + 'ERROR!!!!: ' + 'No se pudo enviar mensaje por el proveedor: ' + str(provider_id) + '  veriique el payload o la ruta de destino')
                f.write('\n' + str(e))

                print("tttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttt")
                print("No se pudo enviar mensaje por el proveedor 1")
                print("tttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttt")
                
                return json.dumps(payload), response.text, 0

    def process_campaign(self):
        
        campaign = self.model.select_all_campaign_agenda_excel_model()
        # for campaign in campaigns:
        if(campaign != None): 
            print("***************************************************")
            global f 
            f = open (str( config('PATH_LOG_CAMPAIGN') + str(campaign[0]) +".log"),'w')
            f.write('***** campaña seleccionada ******')
            f.write('\nid de campana: ' + str(campaign[0]))
            f.write('\n' + 'El usuario tiene el id:' + str(campaign[5]))
            f.write('\n' + 'hora de envio de sms de la campaña:' + str(datetime.now()))

            print(campaign)
            print("***************************************************")
            # print(campaign[0])
            sms_campaign = self.model.select_sms_campaign_by_id( campaign[0])

            is_scheduled = True if campaign[3] != None else False
            # is_excel = True if campaign[6] == 3 else False

            if(is_scheduled):
                f.write('\n' + 'La campaña esta agendado para el dia: ' + str(campaign[3]))
                print(campaign[3])
                is_data_more_than_today = True if datetime.strptime("{}".format(campaign[3]), '%Y-%m-%d %H:%M:%S') > datetime.now() else False
            else :
                f.write('\n' + 'La campana no esta agendada ' )

            if(is_scheduled == False or is_data_more_than_today == True):
                is_excel = True if campaign[6] == 3 else False
                # sms_campaign = self.model.select_sms_campaign_by_id( campaign[0])

                #TODO: CAMBIAR EL STATUS DE CAMPAÑA A PROCESANDO 3
                f.write('\n' + 'Actualizando el estado de la campaña a : 3 ')
                self.model.change_state_campaign(campaign[0], 3)

                if(is_excel):
                    print('leer rows y guardarlos en una variable')
                    excel_sms_campaign = self.model.select_excel_sms_by_id(campaign[0])
                    f.write('\n' + 'Es una campaña creada por excel, el id del excel es: ' + str(excel_sms_campaign[0]))
                    # self.read_excel(r"C:\xampp\htdocs\enviamas9_production\public/storage/ExcelCampaing/Hola.xlsx", sms_campaign,campaign)
                    self.read_excel(config('PUBLIC_PATH_SMS_EXCEL') + excel_sms_campaign[1], sms_campaign,campaign)

                if ( campaign[6] == 2):
                    print('traer contactos de agendas y guardarlos en variable')
                    contacts_by_agenda = self.model.select_all_contact_by_agenda_model(campaign[4])
                    f.write('\n' + 'Es una campana creada por agenda, el id de la agenda es: ' + str(campaign[4]))
                    f.write('\n' + 'La agenda cuenta con un numero de contactos igual a: ' + str(len(contacts_by_agenda)))
                    sms_campaign = self.model.select_sms_campaign_by_id( campaign[0])
                    self.send_sms_by_agenda( campaign, contacts_by_agenda, sms_campaign)
                
                if ( campaign[6] == 1):
                    print('traer contactos de agendas y guardarlos en variable')
                    print('este es el individual!!!!!!!!!!!!!!')

                    contactos = eval(campaign[4])
                    print(contactos)
                    sms_campaign = self.model.select_sms_campaign_by_id( campaign[0])
                    self.send_sms_individuales( campaign, contactos, sms_campaign)

                if(campaign[6] == 4):
                    print('leer json y guardarlos en una variable')
                    json_sms_campaign = self.model.select_json_sms_by_id(campaign[0])
                    print("2222222222222222222222222222222222")
                    print(json_sms_campaign)
                    f.write('\n' + 'Es una campaña creada por json, el id del json es: ' + str(json_sms_campaign[0]))
                    # self.read_excel(r"C:\xampp\htdocs\enviamas9_production\public/storage/ExcelCampaing/Hola.xlsx", sms_campaign,campaign)
                    self.read_json(config('PUBLIC_PATH_SMS_EXCEL') + json_sms_campaign[1], sms_campaign,campaign)

            else:
                print("Esta agendado pero la echa programada ya paso")

            print('****************** -- for')
            # TODO: CAMBIAR EL ESTADO DE LA CAMPANA
            f.write('\n' + 'Actualizando el estado de la campaña a : 1 ')
            self.model.change_state_campaign(campaign[0], 1)

            return self.view.list_campaign(campaign)


# enviarlos datos a la vista
class View(object):

    def list_campaign(self, campaign):
        print(campaign)
        print('ver')


controlador = Controller()
controlador.process_campaign()

